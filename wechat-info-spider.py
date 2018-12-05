import itchat
from openpyxl import Workbook

def get_data():
    itchat.auto_login()

    friends = itchat.get_friends(update=True)

    return friends

def parse_data(data):
    friends = []
    for item in data[1:]:
        friend = {
            'NickName':item['NickName'],
            'RemarkName':item['RemarkName'], #备注名
            'Sex':item['Sex'], # 1 男，2 女
            'Province':item['Province'],
            'City':item['City'],
            'Signature':item['Signature'].replace('\n',' ').replace(',',' '),
            'StarFriend':item['StarFriend'],# 星标好友：1是，0否
            'ContactFlag':item['ContactFlag']# 好友类型及权限：1和3好友，259和33027不让他看我的朋友圈，65539不看他的朋友圈，65795两项设置全禁止
        }
        print(friend)
        friends.append(friend)
    return friends


def print_to_excel():
    friends = parse_data(get_data())
    wb=Workbook()
    ws=[]
    ws.append(wb.create_sheet(title='friends')) #utf8->unicode
    ws[0].append(['序号','昵称','备注名','性别','省份','城市','签名','星标好友','好友类型及权限'])
    count=1
    for ff in friends:
        if ff['Sex'] == "1":
            sex = "男"
        else:
            sex = "女"
        if ff['StarFriend'] == "1":
            StarFriend = "星标好友"
        else:
            StarFriend = "普通好友"
        if ff['ContactFlag'] == "259" or ff['ContactFlag'] == "33027":
            Type = "不让他看我朋友圈"
        elif ff['ContactFlag'] == "65539":
            Type = "不看他朋友圈"
        elif ff['ContactFlag'] == "65795":
            Type = "不让他看我朋友圈；不看他朋友圈"
        else:
            Type = "好友"
        ws[0].append([count,ff['NickName'],ff['RemarkName'],sex,ff['Province'],ff['City'],ff['Signature'],StarFriend,Type])
        count+=1
    save_path='Friends_in_wechat.xlsx'
    wb.save(save_path)

if __name__ == "__main__":
    print_to_excel()